"""
Exportacion de una liquidacion (sueldo mensual) a Excel o PDF.
"""
from io import BytesIO


def _fmt_clp(n: int) -> str:
    return "$" + f"{n:,.0f}".replace(",", ".")


def generar_excel_liquidacion(liq: dict, nombre: str, comisiones: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidacion"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "Liquidacion de sueldo"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Trabajador"
    ws["B2"] = nombre
    ws["A3"] = "Periodo"
    ws["B3"] = liq["periodo"]
    ws["A4"] = "Estado"
    ws["B4"] = liq["estado"]
    for r in (2, 3, 4):
        ws[f"A{r}"].font = bold

    row = 6
    ws.cell(row=row, column=1, value="Categoria").font = header_font
    ws.cell(row=row, column=2, value="Rol").font = header_font
    ws.cell(row=row, column=3, value="Cantidad").font = header_font
    ws.cell(row=row, column=4, value="Monto/Unidad").font = header_font
    ws.cell(row=row, column=5, value="Total").font = header_font
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = header_fill
    row += 1

    for c in comisiones:
        ws.cell(row=row, column=1, value=c["categoria"])
        ws.cell(row=row, column=2, value=c["rol"])
        ws.cell(row=row, column=3, value=c["cantidad"])
        ws.cell(row=row, column=4, value=c["monto_por_unidad"])
        ws.cell(row=row, column=5, value=c["total"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sueldo base").font = bold
    ws.cell(row=row, column=5, value=liq["sueldo_base"]).font = bold
    row += 1
    ws.cell(row=row, column=1, value="Total comisiones").font = bold
    ws.cell(row=row, column=5, value=liq["total_comisiones"]).font = bold
    row += 1
    ws.cell(row=row, column=1, value="Ajustes").font = bold
    ws.cell(row=row, column=5, value=liq["ajustes"]).font = bold
    if liq.get("notas_ajustes"):
        row += 1
        ws.cell(row=row, column=1, value="Detalle ajustes:")
        ws.cell(row=row, column=2, value=liq["notas_ajustes"])
    row += 1
    ws.cell(row=row, column=1, value="TOTAL A PAGAR").font = Font(bold=True, size=12)
    ws.cell(row=row, column=5, value=liq["total"]).font = Font(bold=True, size=12)

    for col, width in zip("ABCDE", (26, 16, 12, 16, 16)):
        ws.column_dimensions[col].width = width
    for r in range(6, row + 1):
        ws.cell(row=r, column=5).number_format = '"$"#,##0'
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="left")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf_liquidacion(liq: dict, nombre: str, comisiones: list[dict]) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle("titulo", parent=styles["Heading1"], fontSize=16)
    normal = styles["Normal"]

    elementos = [
        Paragraph("Liquidacion de sueldo", titulo),
        Spacer(1, 6),
        Paragraph(f"<b>Trabajador:</b> {nombre}", normal),
        Paragraph(f"<b>Periodo:</b> {liq['periodo']}", normal),
        Paragraph(f"<b>Estado:</b> {liq['estado']}", normal),
        Spacer(1, 14),
    ]

    data = [["Categoria", "Rol", "Cantidad", "Monto/Unidad", "Total"]]
    for c in comisiones:
        data.append([
            c["categoria"], c["rol"], str(c["cantidad"]),
            _fmt_clp(c["monto_por_unidad"]), _fmt_clp(c["total"]),
        ])
    tabla = Table(data, colWidths=[140, 70, 60, 90, 90])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    resumen = [
        ["Sueldo base", _fmt_clp(liq["sueldo_base"])],
        ["Total comisiones", _fmt_clp(liq["total_comisiones"])],
        ["Ajustes", _fmt_clp(liq["ajustes"])],
        ["TOTAL A PAGAR", _fmt_clp(liq["total"])],
    ]
    tabla_resumen = Table(resumen, colWidths=[200, 120])
    tabla_resumen.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("FONTSIZE", (0, 3), (-1, 3), 12),
        ("LINEABOVE", (0, 3), (-1, 3), 1, colors.black),
        ("TOPPADDING", (0, 3), (-1, 3), 8),
    ]))
    elementos.append(tabla_resumen)

    if liq.get("notas_ajustes"):
        elementos.append(Spacer(1, 12))
        elementos.append(Paragraph(f"<b>Detalle de ajustes:</b> {liq['notas_ajustes']}", normal))

    doc.build(elementos)
    return buf.getvalue()
